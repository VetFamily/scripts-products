# pip3 install pandas
# pip3 install argparse
# pip3 install xlrd


import argparse
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import pandas as pd
from pathlib import Path
import psycopg2
import glob
from sqlalchemy import create_engine
from sqlalchemy.engine.url import URL
from sqlalchemy.sql import text

from config import config


def getArguments():
    parser = argparse.ArgumentParser(description='Generation du fichier des nouveaux produits')
    # parser.add_argument('-o', '--output', help='Output file name', default='stdout')
    required_args = parser.add_argument_group('required named arguments')
    required_args.add_argument('-d', '--date', help='Dossier contenant les fichiers de logs', required=True)
    return parser.parse_args()


def process():
    df = pd.DataFrame(columns=['Id', 'Dénomination_temp', 'Conditionnement_temp', 'Laboratoire_temp',
                               'Obsolète_temp', 'Invisible_temp', 'ID classe thérapeutique_temp', 'Code GTIN',
                               'Autre code GTIN', 'ID classe thérapeutique', 'Dénomination', 'Conditionnement',
                               'Laboratoire', 'Obsolète', 'Invisible', 'Types', 'Espèces',
                               'Code_Alcyon', 'Dénomination_Alcyon', 'Tarif_Alcyon',
                               'Code_Centravet', 'Dénomination_Centravet', 'Tarif_Centravet',
                               'Code_Coveto', 'Dénomination_Coveto', 'Tarif_Coveto',
                               'Code_Alibon', 'Dénomination_Alibon', 'Tarif_Alibon',
                               'Code_Vetapro', 'Dénomination_Vetapro', 'Tarif_Vetapro',
                               'Code_Vetys', 'Dénomination_Vetys', 'Tarif_Vetys',
                               'Code_Hippocampe', 'Dénomination_Hippocampe', 'Tarif_Hippocampe',
                               'Code_Agripharm', 'Dénomination_Agripharm', 'Tarif_Agripharm',
                               'Code_Elvetis', 'Dénomination_Elvetis', 'Tarif_Elvetis',
                               'Code_Longimpex', 'Dénomination_Longimpex', 'Tarif_Longimpex'])

    if os.path.isfile(logDir + date[0:6] + "_Nouveaux produits_catalogues.xlsx"):
        os.remove(logDir + date[0:6] + "_Nouveaux produits_catalogues.xlsx")

    # Search existing products in database
    query_products = text("""
        select id as produit_id, denomination as denomination_temp, conditionnement as conditionnement_temp, 
            laboratoire_id as laboratoire_id_temp, obsolete as obsolete_temp, invisible as invisible_temp, 
            famille_therapeutique_id as famille_therapeutique_id_temp, code_gtin
        from produits
        where code_gtin is not null""")
    df_products = pd.read_sql_query(query_products, connection)
    df_products['code_gtin'] = pd.to_numeric(df_products['code_gtin'])

    # Search existing therapeutic classes in database
    query_classes = text("""
        select id as famille_therapeutique_id, CONCAT(classe1_code, coalesce(classe2_code, ''), 
            coalesce(classe3_code, '')) as famille_therapeutique
        from familles_therapeutiques""")
    df_classes = pd.read_sql_query(query_classes, connection)

    # Aggregate all files of new products
    for f in sorted(glob.glob(r'' + logDir + 'new_products_*.xlsx')):
        print(f'Processing file "{os.path.basename(f)}" ...')

        df_file = pd.read_excel(f, header=None)
        df_file = df_file.iloc[:, 0:25]
        cent_name = Path(f).stem.split('_')[2]
        if cent_name == 'Alcyon':
            cent_id = 1
        elif cent_name == 'Centravet':
            cent_id = 2
        elif cent_name == 'Coveto':
            cent_id = 3
        elif cent_name == 'Alibon':
            cent_id = 4
        elif cent_name == 'Vetapro':
            cent_id = 5
        elif cent_name == 'Vetys':
            cent_id = 6
        elif cent_name == 'Hippocampe':
            cent_id = 7
        elif cent_name == 'Agripharm':
            cent_id = 8
        elif cent_name == 'Elvetis':
            cent_id = 9
        elif cent_name == 'Longimpex':
            cent_id = 10

        df_file.columns = ['code_distributeur', 'code_gtin', 'code_ean', 'code_amm', 'code_remplacement',
                           'designation', 'categorie', 'famille', 'sous_famille', 'famille_commerciale',
                           'laboratoire', 'statut', 'classe_therapeutique', 'agrement', 'poids_unitaire',
                           'sous_unite_revente', 'gestion_stock', 'taux_tva', 'quantite_tarif',
                           'prix_unitaire_hors_promo', 'prix_unitaire_promo', 'date_debut_promo', 'date_fin_promo',
                           'quantite_ug', 'code_cip']

        temp = pd.merge(df_file[pd.notnull(df_file['code_gtin'])], df_products, on='code_gtin', how='left')
        df_temp = pd.concat([temp, df_file[pd.isnull(df_file['code_gtin'])]], axis=0, sort=False, ignore_index=True)
        del temp

        df_temp.drop(
            ['agrement', 'poids_unitaire', 'sous_unite_revente', 'gestion_stock', 'taux_tva', 'quantite_tarif',
             'prix_unitaire_hors_promo', 'prix_unitaire_promo', 'date_debut_promo', 'date_fin_promo', 'quantite_ug',
             'code_cip'], axis=1, inplace=True)
        df_temp['denomination'] = df_temp['designation']
        df_temp['obsolete'] = df_temp['obsolete_temp']
        df_temp['invisible'] = df_temp['invisible_temp']
        df_temp['conditionnement'] = np.nan

        # Laboratories
        # Search existing laboratories in database
        query_labs = text("""
            select laboratoire_id, nom_laboratoire as laboratoire 
            from centrale_laboratoire
            where laboratoire_id is not null and centrale_id = :id""")
        df_labs = pd.read_sql_query(query_labs, connection, params={'id': cent_id})

        df_temp = pd.merge(df_temp, df_labs, on=['laboratoire'], how='left')
        df_temp.loc[df_temp['laboratoire_id'].isnull(), 'laboratoire_id'] = df_temp['laboratoire']

        # Types
        df_temp = df_temp.replace({'categorie': {'ALI': 'Aliment', 'DIV': 'Divers', 'MAT': 'Matériel',
                                                 'MED': 'Médicament', 'ATB': 'Antibiotique'}})
        df_temp.loc[df_temp['sous_famille'].notnull() & (
                str(df_temp['sous_famille']) == 'ATB'), 'categorie'] = 'Antibiotique'

        # Therapeutic classes
        df_temp['famille_therapeutique'] = df_temp['classe_therapeutique'].str.replace(' ', '').str[:4]
        temp = pd.merge(df_temp[pd.notnull(df_temp['famille_therapeutique'])], df_classes,
                        on=['famille_therapeutique'], how='left')
        df_temp = pd.concat([temp, df_temp[pd.isnull(df_temp['famille_therapeutique'])]], axis=0, sort=False,
                            ignore_index=True)
        del temp

        df_temp = df_temp[
            ['produit_id', 'denomination_temp', 'conditionnement_temp', 'laboratoire_id_temp', 'obsolete_temp',
             'invisible_temp', 'famille_therapeutique_id_temp', 'code_gtin', 'classe_therapeutique',
             'famille_therapeutique_id', 'denomination', 'conditionnement', 'laboratoire_id', 'obsolete',
             'invisible', 'categorie', 'sous_famille', 'code_distributeur', 'designation', 'prix_unitaire_hors_promo']]
        df_temp.columns = ['Id', 'Dénomination_temp', 'Conditionnement_temp', 'Laboratoire_temp', 'Obsolète_temp',
                           'Invisible_temp', 'ID classe thérapeutique_temp', 'Code GTIN', 'Autre code GTIN',
                           'ID classe thérapeutique', 'Dénomination', 'Conditionnement', 'Laboratoire', 'Obsolète',
                           'Invisible', 'Types', 'Espèces', 'Code_' + cent_name, 'Dénomination_' + cent_name,
                           'Tarif_' + cent_name]

        df = pd.concat([df, df_temp.drop_duplicates()], axis=0, sort=False, ignore_index=True)

        # Sort dataframe
        df = df.sort_values(by=['Laboratoire', 'Code GTIN'])

        return df


def create_excel_file(filename, df, append):
    if append:
        wb = load_workbook(filename)
    else:
        wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=(False if append else True)):
        ws.append(r)
    wb.save(filename)


if __name__ == "__main__":
    # Getting args
    args = getArguments()
    date = args.date

    logDir = '/home/ftpusers/amadeo/script-catalogues/' + date + "/"

    print(f'** Generate file of new products of catalogs **')

    try:
        # Reading parameters of database
        params = config()

        # Connecting to the PostgreSQL server
        print('Connecting to the PostgreSQL database...')
        engine = create_engine(URL(**params), echo=False)
        connection = engine.connect()

        df_result = process()

        # Create Excel file
        print("Generating Excel file...")
        create_excel_file(logDir + date[0:6] + "_Nouveaux produits_catalogues.xlsx", df_result, False)

    except (Exception, psycopg2.Error) as error:
        print("Failed : ", error)
        # trans.rollback()
        raise

    finally:
        # closing database connection.
        if connection:
            connection.close()
            print("PostgreSQL connection is closed")
