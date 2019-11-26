# pip3 install pandas
# sudo apt-get build-dep python3-psycopg2
# pip3 install psycopg2-binary
# pip3 install SQLAlchemy
# pip3 install openpyxl
# pip3 install xlrd


import argparse
from datetime import datetime
import os
import pandas as pd
import glob
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import psycopg2
import shutil
from sqlalchemy import create_engine
from sqlalchemy.engine.url import URL
from sqlalchemy.sql import text

from config import config


def getArguments():
    parser = argparse.ArgumentParser(description='Lancement du traitement du fichier de valorisations Vetapro')
    # parser.add_argument('-o', '--output', help='Output file name', default='stdout')
    optional_args = parser.add_argument_group('optional named arguments')
    optional_args.add_argument('-s', '--simulate', help='Simulation de l\'integration du fichier', action='store_true')
    return parser.parse_args()


def insert_new_product(df):
    """Create new product in database

    Parameters:
    row (DataFrame row) : row of products Dataframe

    Returns:
    int : ID of created product"""

    global count_of_new_products

    df_temp = df[df['Id'].isnull()].copy()

    df_temp.drop(df_temp.columns.difference(
        ['Code GTIN', 'Autre code GTIN', 'ID classe thérapeutique', 'Dénomination', 'Conditionnement',
         'Laboratoire', 'Obsolète', 'Invisible']), axis=1, inplace=True)
    df_temp = df_temp.drop_duplicates()
    df_temp.columns = ['code_gtin', 'code_gtin_autre', 'famille_therapeutique_id', 'denomination',
                       'conditionnement', 'laboratoire_id', 'obsolete', 'invisible']
    # df_temp.loc[df_temp['Code GTIN'].notnull() & (str(df_temp['Code GTIN']) == ''), 'Code GTIN'] = np.nan
    df_temp.to_sql('produits', engine, if_exists='append', index=False, chunksize=1000)
    count_of_new_products = len(df_temp.index)

    del df_temp


def insert_types(df):
    global df_logs, count_of_types

    df_temp = df[df['Id'].isnull()].copy()
    df_temp.drop(df_temp.columns.difference(['produit_id', 'Types']), axis=1, inplace=True)

    # Log products without types
    df_without_types = df_temp.loc[df_temp['Types'].isnull(), :]
    for index, row in df_without_types.iterrows():
        df_logs = df_logs.append(
            pd.DataFrame([[os.path.basename(f), row['produit_id'], 'Erreur', 'Pas de type renseigné']
                          ]),
            ignore_index=True)

    df_temp = df_temp.loc[df_temp['Types'].notnull(), :]

    if len(df_temp.index) > 0:
        df_types = pd.DataFrame(df_temp['Types'].str.split('|').tolist(), index=df_temp['produit_id']).stack()
        df_types = df_types.reset_index()[[0, 'produit_id']]
        df_types.columns = ['type_id', 'produit_id']
        df_types = df_types.replace(
            {'type_id': {"Aliment": 1, "Médicament": 2, "Antibiotique": 3, "Divers": 4, "Matériel": 5}})
        df_types.to_sql('produit_type', engine, if_exists='append', index=False, chunksize=1000)
        count_of_types = len(df_types.index)

        del df_types

    del df_temp


def update_product(df):
    global count_of_upd_products

    df_temp = df[df['Id'].notnull() & (df['Code GTIN'].notnull() | df['Autre code GTIN'].notnull())].copy()

    for index, row in df_temp.iterrows():
        update = False

        if row['Code GTIN'] is not None and pd.notna(row['Code GTIN']):
            query = text("""UPDATE produits SET code_gtin = :code_gtin WHERE id = :id""")
            res = connection.execute(query, code_gtin=row['Code GTIN'], id=row['Id'])
            if res.rowcount == 1:
                update = True

        if row['Autre code GTIN'] is not None and pd.notna(row['Autre code GTIN']):
            query = text("""UPDATE produits SET code_gtin_autre = :code_gtin WHERE id = :id""")
            res = connection.execute(query, code_gtin=row['Autre code GTIN'], id=row['Id'])
            if res.rowcount == 1:
                update = True

        if update:
            count_of_upd_products += 1

    del df_temp


def insert_species(df):
    global df_logs, count_of_species

    df_temp = df[df['Id'].isnull()].copy()
    df_temp.drop(df_temp.columns.difference(['produit_id', 'Espèces']), axis=1, inplace=True)

    # Log products without species
    df_without_species = df_temp.loc[df_temp['Espèces'].isnull(), :]
    for index, row in df_without_species.iterrows():
        df_logs = df_logs.append(
            pd.DataFrame([[os.path.basename(f), row['produit_id'], 'Erreur', 'Pas d\'espèce renseignée']
                          ]),
            ignore_index=True)

    df_temp = df_temp.loc[df_temp['Espèces'].notnull(), :]

    if len(df_temp.index) > 0:
        df_species = pd.DataFrame(df_temp['Espèces'].str.split('|').tolist(), index=df_temp['produit_id']).stack()
        df_species = df_species.reset_index()[[0, 'produit_id']]
        df_species.columns = ['espece_id', 'produit_id']
        df_species = df_species.replace(
            {'espece_id': {"Canine": 1, "Equine":2, "Rurale": 3, "Porc": 4, "Volaille": 5, "Autres": 6}})
        df_species.to_sql('espece_produit', engine, if_exists='append', index=False, chunksize=1000)
        count_of_species = len(df_species.index)

        del df_species

    del df_temp


def insert_central_codes(df, cent_id, cent_name):
    global df_logs

    count = 0
    date = pd.to_datetime("2016-01-01", format='%Y-%m-%d', errors='coerce')

    df_temp = df[df['Code_' + cent_name].notnull()].copy()
    df_temp.drop(df_temp.columns.difference(
        ['produit_id', 'Code_' + cent_name, 'Dénomination_' + cent_name, 'Tarif_' + cent_name]), axis=1, inplace=True)
    df_temp.columns = ['code_produit', 'nom', 'prix_unitaire_hors_promo', 'produit_id']
    df_temp['code_produit'] = df_temp['code_produit'].dropna().apply(lambda x: str(x))
    df_temp['nom'] = df_temp['nom'].dropna().apply(lambda x: str(x))

    query_cp = text("""
                    SELECT id as centrale_produit_id, code_produit
                    FROM centrale_produit
                    WHERE centrale_id = :id""")
    df_temp = pd.merge(df_temp, pd.read_sql_query(query_cp, connection, params={'id': cent_id}), on='code_produit',
                       how='left')

    df_temp_upd = df_temp.loc[df_temp['centrale_produit_id'].notnull(), :].copy()

    for index, row in df_temp_upd.iterrows():
        # Update code because it already exists
        query_upd = text("""UPDATE centrale_produit SET produit_id = :prodId WHERE id = :id""")
        res_upd = connection.execute(query_upd, id=row['centrale_produit_id'], prodId=row['produit_id'])
        if res_upd.rowcount != 1:
            df_logs = df_logs.append(pd.DataFrame(
                [[os.path.basename(f), row['produit_id'], 'Erreur', 'Code ' + cent_name + ' non mis à jour']]),
                                     ignore_index=True)
        else:
            count += 1

    df_temp_new = df_temp.loc[df_temp['centrale_produit_id'].isnull(), :].copy()

    # Insert into centrale_produit
    df_temp_new_cp = df_temp_new[['produit_id', 'code_produit']]
    df_temp_new_cp.insert(0, 'centrale_id', cent_id)
    df_temp_new_cp.to_sql('centrale_produit', engine, if_exists='append', index=False, chunksize=1000)
    count += len(df_temp_new_cp.index)

    df_temp_new = df_temp_new.drop('centrale_produit_id', axis=1)
    df_temp_new = pd.merge(df_temp_new, pd.read_sql_query(query_cp, connection, params={'id': cent_id}),
                           on='code_produit', how='left')

    if len(df_temp_new.index) > 0:
        # Insert into centrale_produit_denominations
        df_temp_new_cpd = df_temp_new[['centrale_produit_id', 'nom']]
        df_temp_new_cpd.insert(0, 'date_creation', date)
        df_temp_new_cpd.to_sql('centrale_produit_denominations', engine, if_exists='append', index=False, chunksize=1000)

        del df_temp_new_cpd

    count_of_centrals_codes[cent_name] = count

    del df_temp, df_temp_upd, df_temp_new


def process():
    df = df_init.where(pd.notnull(df_init), None)
    df['Obsolète'] = df['Obsolète'].astype(bool)
    df['Invisible'] = df['Invisible'].astype(bool)
    df['Id'] = pd.to_numeric(df['Id'])
    df['Code GTIN'] = df['Code GTIN'].dropna().apply(lambda x: str(int(x)))
    df['Dénomination'] = df['Dénomination'].dropna().apply(lambda x: str(x))
    df['Conditionnement'] = df['Conditionnement'].dropna().apply(lambda x: str(x))
    for col in df.columns:
        if 'Code_' in col:
            df[col] = df[col].dropna().apply(lambda x: str(int(x)) if type(x) is float else x)

    # Insert new products
    insert_new_product(df)

    # Add products IDs
    query_products = text("""
                SELECT id as produit_id, denomination, conditionnement, laboratoire_id, obsolete, invisible, code_gtin 
                FROM produits""")
    df_products = pd.read_sql_query(query_products, connection)
    temp = pd.merge(df[pd.isnull(df['Id'])], df_products,
                    left_on=['Code GTIN', 'Dénomination', 'Conditionnement', 'Laboratoire', 'Obsolète', 'Invisible'],
                    right_on=['code_gtin', 'denomination', 'conditionnement', 'laboratoire_id', 'obsolete',
                              'invisible'], how='left')
    df = pd.concat([temp, df[pd.notnull(df['Id'])]], axis=0, sort=False, ignore_index=True)
    df.loc[df['produit_id'].isnull(), 'produit_id'] = df['Id']
    df['produit_id'] = pd.to_numeric(df['produit_id'])

    # Add Vetapro codes
    df.loc[(df['Id'].isnull() & df['Code_Vetapro'].isnull()), 'Code_Vetapro'] = df['produit_id']

    # Insert types of new products
    insert_types(df)

    # Insert species of new products
    insert_species(df)

    # Update products
    update_product(df)

    # Insert Alcyon codes
    insert_central_codes(df, 1, 'Alcyon')

    # Insert Centravet codes
    insert_central_codes(df, 2, 'Centravet')

    # Insert Coveto codes
    insert_central_codes(df, 3, 'Coveto')

    # Insert Alibon codes
    insert_central_codes(df, 4, 'Alibon')

    # Insert Vetapro codes
    insert_central_codes(df, 5, 'Vetapro')

    # Insert Vetys codes
    insert_central_codes(df, 6, 'Vetys')

    # Insert Hippocampe codes
    insert_central_codes(df, 7, 'Hippocampe')

    # Insert Agripharm codes
    insert_central_codes(df, 8, 'Agripharm')

    # Insert Elvetis codes
    insert_central_codes(df, 9, 'Elvetis')

    # Insert Longimpex codes
    insert_central_codes(df, 10, 'Longimpex')


def create_excel_file(filename, df, append):
    if append:
        wb = load_workbook(filename)
    else:
        wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=(False if append else True)):
        ws.append(r)
    wb.save(filename)


if __name__ == "__main__":
    print(f'** Processing new products **')

    # Getting args
    args = getArguments()
    simulation = args.simulate

    initDir = './fichiers/'
    workDir = './encours/'
    historicDir = './historiques/nouveaux/' + datetime.now().strftime('%Y%m%d') + "/"
    logDir = './logs/nouveaux/' + datetime.now().strftime('%Y%m%d') + "/"

    # Create directories if not exist
    os.makedirs(workDir, exist_ok=True)
    os.makedirs(historicDir, exist_ok=True)
    os.makedirs(logDir, exist_ok=True)

    try:
        # Reading parameters of database
        params = config(section="postgresql")

        # Connecting to the PostgreSQL server
        print('Connecting to the PostgreSQL database...')
        engine = create_engine(URL(**params), echo=False)
        connection = engine.connect()

        # Create empty logs dataframes
        df_logs = pd.DataFrame(columns=['Fichier', 'ID Produit', 'Type', 'Description'])
        if not os.path.isfile(logDir + "products_errors.xlsx"):
            create_excel_file(logDir + "products_errors.xlsx", df_logs, False)

        for f in glob.glob(r'' + initDir + '/*.[xX][lL][sS][xX]'):
            print(f'Processing file "{os.path.basename(f)}" ...')

            # Begin transaction
            trans = connection.begin()

            # Move file to working directory
            shutil.move(f, workDir + os.path.basename(f))

            if os.stat(workDir + os.path.basename(f)).st_size > 0:
                # Initialize counts
                count_of_new_products = 0
                count_of_types = 0
                count_of_species = 0
                count_of_upd_products = 0
                count_of_centrals_codes = {"Alcyon": 0, "Centravet": 0, "Coveto": 0, "Alibon": 0, "Vetapro": 0,
                                           "Vetys": 0, "Hippocampe": 0, "Agripharm": 0, "Elvetis": 0, "Longimpex": 0}
                # Read Excel file
                df_init = pd.read_excel(workDir + os.path.basename(f))
                # Process file
                process()
            else:
                print(f'File "{os.path.basename(f)}" is empty !')

            # Move source file
            shutil.move(workDir + os.path.basename(f), historicDir + os.path.basename(f))

            trans.commit()
            print(f'File "{os.path.basename(f)}" is treated :')
            print(f'\t{count_of_new_products} products created')
            print(f'\t{count_of_types} types created')
            print(f'\t{count_of_species} species created')
            print(f'\t{count_of_upd_products} products updated')
            for cent in count_of_centrals_codes:
                print(f'\t{count_of_centrals_codes[cent]} codes created or updated for {cent}')

        # Create Excel file of products errors
        create_excel_file(logDir + "products_errors.xlsx", df_logs.drop_duplicates(), True)

    except (Exception, psycopg2.Error) as error:
        if connection:
            print("Error : ", error)

    finally:
        # closing database connection.
        if connection:
            connection.close()
            print("PostgreSQL connection is closed")
