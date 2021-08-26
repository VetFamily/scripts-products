# pip3 install pandas
# sudo apt-get build-dep python3-psycopg2
# pip3 install psycopg2-binary
# pip3 install SQLAlchemy
# pip3 install openpyxl
# pip3 install xlrd


import argparse
import glob
import os
import shutil
from datetime import datetime

import numpy as np
import pandas as pd
import psycopg2
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import create_engine
from sqlalchemy.engine.url import URL
from sqlalchemy.sql import text

from config import config
from src.common import common, constant


def getArguments():
    parser = argparse.ArgumentParser(description='Lancement du traitement du fichier de valorisations Vetapro')
    # parser.add_argument('-o', '--output', help='Output file name', default='stdout')
    required_args = parser.add_argument_group('required named arguments')
    required_args.add_argument('-c', '--country', help='ID of country', required=True)
    return parser.parse_args()


def check_values(values, ref):
    """
    Check that all values are in predefined list 'ref'.
    Parameters
    ----------
    values : named pandas Series
    ref : list (or pandas Series)
    """
    diff = set(values) - set(ref)
    if len(diff) > 0:
        raise ValueError(f"Unexpected value(s) in column {values.name}: {diff}")


def check_dataframe(df, con):
    """
    Perform multiple checks on dataframe originating
    from generate_from_puchase_logs.py
    Parameters
    ----------
    df : pandas dataframe
    con : connection to the postgreSQL database
    """
    df = df[df['Id'].isnull()]

    if len(df) > 0:
        # check that all rows have name, packaging and supplier
        for col in ['Dénomination', 'Laboratoire', 'Types', 'Espèces']:
            if df[col].isnull().values.any():
                raise ValueError(f"Empty value(s) in column {col}")

        # check that all values in 'Types' column exist in table 'types'
        types = pd.read_sql_query("select code as nom from types where obsolete is false", con)
        check_values(df['Types'], types['nom'])

        # check that all values in 'Espèces' column exist in table 'especes'
        especes = pd.read_sql_query("select nom from especes where obsolete is false", con)
        df_series = df['Espèces'].str.split('|').apply(pd.Series).stack()
        df_series.name = 'nom'
        check_values(df_series, especes['nom'])

        # check that all values in columns 'Obsolète' and 'Invisible' are either 'True' or 'False'
        check_values(df['Obsolète'], ['True', 'False'])
        check_values(df['Invisible'], ['True', 'False'])

        # check that all values in 'ID classe thérapeutique' column exist in table
        #   'familles_therapeutiques'
        familles = pd.read_sql_query(
            "select id from familles_therapeutiques where obsolete is false",
            con
        )
        check_values(pd.to_numeric(df['ID classe thérapeutique']), familles['id'])


def insert_new_product(df):
    global count_of_new_products
    col_names = ['code_gtin', 'code_gtin_autre', 'famille_therapeutique_id', 'denomination',
                 'conditionnement', 'laboratoire_id', 'obsolete', 'invisible']
    col_labels = ['Code GTIN', 'Autre code GTIN', 'ID classe thérapeutique', 'Dénomination',
                  'Conditionnement', 'Laboratoire', 'Obsolète', 'Invisible']

    # step 1: update df['Id'] when same product found in 'produits' table

    produits = pd.read_sql_query(
        """select id, code_gtin, code_gtin_autre, famille_therapeutique_id, denomination,
                  conditionnement, laboratoire_id, obsolete, invisible
           from produits""",
        connection
    )

    df = pd.merge(
        df,
        produits,
        how='left',
        left_on=col_labels,
        right_on=col_names
    )
    df.loc[df['Id'].isnull(), 'Id'] = df['id']
    del df['id']

    # step 2: insert new products

    df_temp = df[df['Id'].isnull()].copy()

    df_temp = df_temp[col_labels]
    df_temp.drop_duplicates(inplace=True)
    df_temp.columns = col_names
    # df_temp.loc[df_temp['Code GTIN'].notnull() & (str(df_temp['Code GTIN']) == ''), 'Code GTIN'] = np.nan
    df_temp.to_sql('produits', engine, if_exists='append', index=False, chunksize=1000)
    count_of_new_products = len(df_temp)

    del df_temp

    return df


def insert_types(df):
    global df_logs, count_of_types

    df_temp = df[df['Id'].isnull()].copy()
    df_temp = df_temp[['produit_id', 'Types']]

    # Log products without types
    df_without_types = df_temp[df_temp['Types'].isnull()]
    for index, row in df_without_types.iterrows():
        df_logs = df_logs.append(
            pd.DataFrame([[os.path.basename(f), row['produit_id'], 'Erreur', 'Pas de type renseigné']
                          ]),
            ignore_index=True)

    df_temp = df_temp[df_temp['Types'].notnull()]

    if len(df_temp) > 0:
        df_types = pd.DataFrame(df_temp['Types'].str.split('|').tolist(), index=df_temp['produit_id']).stack()
        df_types = df_types.reset_index()[[0, 'produit_id']]
        df_types.columns = ['type_id', 'produit_id']
        df_types = df_types.replace(
            {'type_id': {"ALI": 1, "ATB": 2, "DIV": 3, "MAT": 4, "MED": 5, "BIO": 6}})
        df_types.drop_duplicates(inplace=True)
        df_types.to_sql('produit_type', engine, if_exists='append', index=False, chunksize=1000)
        count_of_types = len(df_types)

        del df_types

    del df_temp


def update_product(df):
    global count_of_upd_products

    df_temp = df[df['Id'].notnull() & (df['Code GTIN'].notnull() | df['Autre code GTIN'].notnull())].copy()

    for index, row in df_temp.iterrows():
        update = False

        if pd.notnull(row['Code GTIN']):
            query = text("""UPDATE produits SET code_gtin = :code_gtin WHERE id = :id""")
            res = connection.execute(query, code_gtin=row['Code GTIN'], id=row['Id'])
            if res.rowcount == 1:
                update = True

        if pd.notnull(row['Autre code GTIN']):
            query = text("""UPDATE produits SET code_gtin_autre = :code_gtin WHERE id = :id""")
            res = connection.execute(query, code_gtin=row['Autre code GTIN'], id=row['Id'])
            if res.rowcount == 1:
                update = True

        if update:
            count_of_upd_products += 1

    del df_temp


def insert_species(df):
    global df_logs, count_of_species

    df_temp = df[df['Id'].isnull()].copy()
    df_temp = df_temp[['produit_id', 'Espèces']]

    # Log products without species
    df_without_species = df_temp[df_temp['Espèces'].isnull()]
    for index, row in df_without_species.iterrows():
        df_logs = df_logs.append(
            pd.DataFrame([[os.path.basename(f), row['produit_id'], 'Erreur', 'Pas d\'espèce renseignée']
                          ]),
            ignore_index=True)

    df_temp = df_temp[df_temp['Espèces'].notnull()]

    if len(df_temp) > 0:
        df_species = pd.DataFrame(df_temp['Espèces'].str.split('|').tolist(), index=df_temp['produit_id']).stack()
        df_species = df_species.reset_index()[[0, 'produit_id']]
        df_species.columns = ['espece_id', 'produit_id']
        df_species = df_species.replace(
            {'espece_id': {"Canine": 1, "Equine": 2, "Rurale": 3, "Porc": 4, "Volaille": 5, "Autres": 6}})
        df_species.drop_duplicates(inplace=True)
        df_species.to_sql('espece_produit', engine, if_exists='append', index=False, chunksize=1000)
        count_of_species = len(df_species)

        del df_species

    del df_temp


def insert_central_codes(df, cent_id, cent_name):
    global df_logs

    # stop function early if necessary
    if 'Code_' + cent_name not in df.columns:
        return

    count = 0
    date = pd.to_datetime("2016-01-01", format='%Y-%m-%d', errors='coerce')

    df_temp = df[df['Code_' + cent_name].notnull()].copy()

    labels_dict = {
        'Code_' + cent_name: 'code_produit',
        'Dénomination_' + cent_name: 'nom',
        'Condition_commerciale_' + cent_name: 'cirrina_pricing_condition_id'
    }

    if cent_id in [constant.SOURCE_CIRRINA_ID, constant.SOURCE_SERVIPHAR_ID]:
        df_temp = df_temp[['produit_id', 'Code_' + cent_name, 'Dénomination_' + cent_name,
                           'Condition_commerciale_' + cent_name]]
    else:
        df_temp = df_temp[['produit_id', 'Code_' + cent_name, 'Dénomination_' + cent_name]]
        df_temp['cirrina_pricing_condition_id'] = np.nan

    df_temp.rename(columns=labels_dict, inplace=True)
    df_temp['code_produit'] = df_temp['code_produit'].dropna().astype(str)
    df_temp['nom'] = df_temp['nom'].dropna().astype(str)
    df_temp['cirrina_pricing_condition_id'] = pd.to_numeric(df_temp['cirrina_pricing_condition_id'])

    query_cp = text("""
                        SELECT id as centrale_produit_id, code_produit, cirrina_pricing_condition_id
                        FROM centrale_produit
                        WHERE centrale_id = :sourceId
                        AND country_id = :countryId
                                                        """)
    df_cp = pd.read_sql_query(query_cp, connection, params={'sourceId': cent_id, 'countryId': country_id})
    df_cp['code_produit'] = df_cp['code_produit'].dropna().astype(str)
    df_cp['cirrina_pricing_condition_id'] = pd.to_numeric(df_cp['cirrina_pricing_condition_id'])
    df_temp = pd.merge(df_temp, df_cp, on=['code_produit', 'cirrina_pricing_condition_id'], how='left')
    del df_cp

    if country_id != 1:
        insert_product_country_label(df_temp, False)

    df_temp_upd = df_temp[df_temp['centrale_produit_id'].notnull()].copy()

    for index, row in df_temp_upd.iterrows():
        # Update code because it already exists
        query_upd = text("""UPDATE centrale_produit SET produit_id = :prodId WHERE id = :id""")
        res_upd = connection.execute(query_upd, id=row['centrale_produit_id'], prodId=row['produit_id'])
        if res_upd.rowcount != 1:
            df_logs = df_logs.append(pd.DataFrame(
                [[os.path.basename(f), row['produit_id'], 'Erreur', 'Code ' + cent_name + ' non mis à jour']]),
                                     ignore_index=True)
        else:
            count += 1

    df_temp_new = df_temp[df_temp['centrale_produit_id'].isnull()].copy()

    # Insert into centrale_produit
    df_temp_new_cp = df_temp_new[['produit_id', 'code_produit', 'cirrina_pricing_condition_id']].copy()
    df_temp_new_cp['centrale_id'] = cent_id
    df_temp_new_cp['country_id'] = country_id
    df_temp_new_cp.drop_duplicates(inplace=True)
    df_temp_new_cp.to_sql('centrale_produit', engine, if_exists='append', index=False, chunksize=1000)
    count += len(df_temp_new_cp)

    del df_temp_new['centrale_produit_id']
    df_cp = pd.read_sql_query(query_cp, connection, params={'sourceId': cent_id, 'countryId': country_id})
    df_cp['code_produit'] = df_cp['code_produit'].dropna().astype(str)
    df_cp['cirrina_pricing_condition_id'] = pd.to_numeric(df_cp['cirrina_pricing_condition_id'])
    df_temp_new = pd.merge(df_temp_new, df_cp, on=['code_produit', 'cirrina_pricing_condition_id'], how='left')
    del df_cp

    if len(df_temp_new) > 0:
        # Insert into centrale_produit_denominations
        df_temp_new_cpd = df_temp_new[['centrale_produit_id', 'nom']].copy()
        df_temp_new_cpd['date_creation'] = date
        df_temp_new_cpd.drop_duplicates(inplace=True)
        df_temp_new_cpd.to_sql('centrale_produit_denominations', engine, if_exists='append', index=False,
                               chunksize=1000)

        del df_temp_new_cpd

    count_of_centrals_codes[cent_name] = count

    del df_temp, df_temp_upd, df_temp_new


def insert_product_country(df):
    query = text("""
                    select distinct product_id from product_country
                    where country_id = :countryId
                    """)
    df_tmp = pd.read_sql_query(query, connection, params={'countryId': country_id})
    # Formatting columns
    df_tmp['product_id'] = pd.to_numeric(df_tmp['product_id'])

    df_products = df[~df['produit_id'].isin(df_tmp['product_id'])].copy()
    if len(df_products) > 0:
        df_products = df_products[['produit_id']]
        df_products.rename(columns={'produit_id': 'product_id'}, inplace=True)
        df_products['country_id'] = country_id
        df_products.drop_duplicates(inplace=True)
        df_products.to_sql('product_country', engine, if_exists='append', index=False, chunksize=1000)

    del df_products, df_tmp


def insert_product_country_label(df, is_france):
    query = text("""
                        select distinct l.code as label_code, language_id
                        from label l
                        join country c on c.default_language_id = l.language_id
                        where l.code like 'PROD%'
                """)
    df_tmp = pd.read_sql_query(query, connection, params={'countryId': country_id})

    df_products_labels = df.copy()
    if len(df_products_labels) > 0:
        query_language = text("""select default_language_id from country where id = :countryId""")
        language_id = connection.execute(query_language, countryId=country_id).first()[0]

        if is_france:
            product_name = 'Dénomination'
            product_packaging = 'Conditionnement'
            df_products_labels.loc[df_products_labels[product_packaging].isnull(), product_packaging] = ''
        else:
            product_name = 'nom'
            product_packaging = 'product_packaging'
            df_products_labels[product_packaging] = ''

        df_products_labels['name_label'] = df_products_labels.agg(lambda x: f"PROD{str(int(x['produit_id']))}N", axis=1)
        df_products_labels['packaging_label'] = df_products_labels.agg(lambda x: f"PROD{str(int(x['produit_id']))}P",
                                                                       axis=1)
        df_products_labels['language_id'] = language_id

        # Label of name of product
        df_products_labels_name = df_products_labels[~df_products_labels['name_label'].isin(df_tmp['label_code']) &
                                                     df_products_labels[product_name].notnull()]

        if len(df_products_labels_name) > 0:
            df_products_labels_name = df_products_labels_name[['name_label', product_name, 'language_id']]
            df_products_labels_name.drop_duplicates(subset='name_label', keep='first', inplace=True)
            df_products_labels_name.rename(columns={'name_label': 'code', product_name: 'value'}, inplace=True)
            df_products_labels_name.drop_duplicates(inplace=True)
            df_products_labels_name.to_sql('label', engine, if_exists='append', index=False, chunksize=1000)

        # Label of packaging of product
        df_products_labels_pack = df_products_labels[~df_products_labels['packaging_label'].isin(
            df_tmp['label_code']) & df_products_labels[product_packaging].notnull()]

        if len(df_products_labels_pack) > 0:
            df_products_labels_pack = df_products_labels_pack[['packaging_label', product_packaging, 'language_id']]
            df_products_labels_pack.drop_duplicates(subset='packaging_label', keep='first', inplace=True)
            df_products_labels_pack.rename(columns={'packaging_label': 'code', product_packaging: 'value'}, inplace=True)
            df_products_labels_pack.drop_duplicates(inplace=True)
            df_products_labels_pack.to_sql('label', engine, if_exists='append', index=False, chunksize=1000)

    del df_products_labels, df_tmp


def process():
    df = df_init.where(df_init.notnull(), None)
    df['Obsolète'] = df['Obsolète'] == 'True'
    df['Invisible'] = df['Invisible'] == 'True'
    df['Id'] = pd.to_numeric(df['Id'])
    df['Laboratoire'] = pd.to_numeric(df['Laboratoire'])
    df['Code GTIN'] = df['Code GTIN'].dropna().apply(lambda x: str(int(x)) if type(x) is float else str(x))
    df['Autre code GTIN'] = df['Autre code GTIN'].dropna().apply(lambda x: str(int(x)) if type(x) is float else str(x))
    df['Dénomination'] = df['Dénomination'].dropna().astype(str)
    df['Conditionnement'] = df['Conditionnement'].dropna().astype(str)
    df['ID classe thérapeutique'] = pd.to_numeric(df['ID classe thérapeutique'])

    # Insert new products
    df = insert_new_product(df)

    # Add products IDs
    query_products = text("""
                SELECT id as produit_id, denomination, conditionnement, laboratoire_id, obsolete, invisible, code_gtin
                FROM produits""")
    df_products = pd.read_sql_query(query_products, connection)
    df_products['produit_id'] = pd.to_numeric(df_products['produit_id'])
    df_products['laboratoire_id'] = pd.to_numeric(df_products['laboratoire_id'])

    temp = pd.merge(df[df['Id'].isnull()], df_products,
                    left_on=['Code GTIN', 'Dénomination', 'Conditionnement', 'Laboratoire', 'Obsolète', 'Invisible'],
                    right_on=['code_gtin', 'denomination', 'conditionnement', 'laboratoire_id', 'obsolete',
                              'invisible'], how='left')
    df = pd.concat([temp, df[df['Id'].notnull()]], axis=0, sort=False, ignore_index=True)
    df.loc[df['produit_id'].isnull(), 'produit_id'] = df['Id']
    df['produit_id'] = pd.to_numeric(df['produit_id'])

    # Add Vetapro codes
    if country_id == 1:
        df.loc[df['Id'].isnull(), 'Code_Vetapro'] = df['produit_id']
        df.loc[df['Id'].isnull(), 'Dénomination_Vetapro'] = df[['Dénomination', 'Conditionnement']].apply(
            lambda x: ' '.join(x.dropna()), axis=1)

    # Add country ID for products
    insert_product_country(df)

    if country_id == 1:
        insert_product_country_label(df, True)

    # Remove .0 from codes columns (not for Cirrina, Serviphar and Distrivet)
    for col in df.columns:
        if 'Code_' in col:
            if col in ['Code_Cirrina', 'Code_Serviphar', 'Code_Distrivet']:
                df[col] = df[col].dropna().astype(str)
            else:
                df[col] = df[col].dropna().apply(lambda x: str(int(x)) if type(x) is float else x)

    # Insert types of new products
    insert_types(df)

    # Insert species of new products
    insert_species(df)

    # Update products
    update_product(df)

    # insert codes for each central
    centrals = pd.read_sql_query(
        "select id, nom from centrales where obsolete is false",
        connection
    )
    for index, row in centrals.iterrows():
        insert_central_codes(df, row['id'], row['nom'].capitalize())


def create_excel_file(filename, df, append):
    if append:
        wb = load_workbook(filename)
    else:
        wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=(not append)):
        ws.append(r)
    wb.save(filename)


if __name__ == "__main__":
    try:
        # Reading parameters of database
        params = config(section="postgresql")

        # Connecting to the PostgreSQL server
        print('Connecting to the PostgreSQL database...')
        engine = create_engine(URL(**params), echo=False)
        connection = engine.connect()

        print('** Processing new products **')

        # Getting args
        args = getArguments()
        country_id = int(args.country)
        country_name = common.get_name_of_country(connection, country_id)

        now = datetime.now().strftime('%Y%m%d')

        initDir = os.path.join('fichiers', 'nouveaux', country_name)
        workDir = os.path.join('encours', 'nouveaux', country_name)
        historicDir = os.path.join('historiques', 'nouveaux', country_name, now)
        logDir = os.path.join('logs', 'nouveaux', country_name, now)

        # Create directories if not exist
        os.makedirs(workDir, exist_ok=True)
        os.makedirs(historicDir, exist_ok=True)

        # Create empty logs dataframes
        df_logs = pd.DataFrame(columns=['Fichier', 'ID Produit', 'Type', 'Description'])
        errors_file = os.path.join(logDir, "products_errors.xlsx")

        for f in glob.glob(os.path.join(initDir, '*.[xX][lL][sS][xX]')):
            print(f'Processing file "{os.path.basename(f)}" ...')

            # Begin transaction
            trans = connection.begin()

            # Move file to working directory
            shutil.move(f, os.path.join(workDir, os.path.basename(f)))

            # Initialize counts
            count_of_new_products = 0
            count_of_types = 0
            count_of_species = 0
            count_of_upd_products = 0
            count_of_centrals_codes = {"Alcyon": 0, "Centravet": 0, "Coveto": 0, "Alibon": 0, "Vetapro": 0,
                                       "Vetys": 0, "Hippocampe": 0, "Agripharm": 0, "Elvetis": 0, "Longimpex": 0,
                                       "Direct": 0, "Cedivet": 0, "Covetrus": 0, "Apoex": 0, "Kruuse": 0,
                                       "Apotek1": 0, "Cirrina": 0, "Serviphar": 0, "Soleomed": 0, "Veso": 0,
                                       "Distrivet": 0, "Elasa": 0, "Centauro": 0
                                       }
            if os.stat(os.path.join(workDir, os.path.basename(f))).st_size > 0:
                # Read Excel file
                df_init = pd.read_excel(os.path.join(workDir, os.path.basename(f)), dtype=str)
                # Check file
                check_dataframe(df_init, connection)
                # Process file
                process()
            else:
                print(f'File "{os.path.basename(f)}" is empty !')

            # Move source file
            shutil.move(os.path.join(workDir, os.path.basename(f)),
                        os.path.join(historicDir, os.path.basename(f)))

            trans.commit()
            print(f'File "{os.path.basename(f)}" is treated :')
            print(f'\t{count_of_new_products} products created')
            print(f'\t{count_of_types} types created')
            print(f'\t{count_of_species} species created')
            print(f'\t{count_of_upd_products} products updated')
            for cent in count_of_centrals_codes:
                print(f'\t{count_of_centrals_codes[cent]} codes created or updated for {cent}')

        # Create Excel file of products errors
        if len(df_logs) > 0:
            os.makedirs(logDir, exist_ok=True)
            create_excel_file(os.path.join(logDir, "products_errors.xlsx"),
                              df_logs.drop_duplicates(), append=False)
            print('WARNING : there are an error file !')

    except (Exception, psycopg2.Error) as error:
        if connection:
            print("Error : ", error)

    finally:
        # closing database connection.
        if connection:
            connection.close()
            print("PostgreSQL connection is closed")
