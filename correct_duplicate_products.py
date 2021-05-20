# pip3 install pandas
# sudo apt-get build-dep python3-psycopg2
# pip3 install psycopg2-binary
# pip3 install SQLAlchemy
# pip3 install openpyxl
# pip3 install xlrd
import argparse
from datetime import datetime
import os
import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import psycopg2
import shutil
from sqlalchemy import create_engine
from sqlalchemy.engine.url import URL
from sqlalchemy.sql import text

from config import config
from src.common import common, constant


def getArguments():
    parser = argparse.ArgumentParser(description='Lancement du traitement des doublons')
    # parser.add_argument('-o', '--output', help='Output file name', default='stdout')
    required_args = parser.add_argument_group('required named arguments')
    required_args.add_argument('-c', '--country', help='ID of country', required=True)
    return parser.parse_args()


def process():
    global df_logs

    if country_id == constant.COUNTRY_FRANCE_ID:
        clients = {"bourgelat": 0, "vetoavenir": 0, "vetapharma": 0, "vetharmonie": 0, "cristal": 0,
                   "symbioveto": 0, "clubvet": 0, "vetodistribution": 0, "vetfamily": 0, "vetapro": 0}
    else:
        clients = {"vetfamily": 0}

    # Getting country of products
    query = text("""
                    select distinct product_id from product_country
                    where country_id = :countryId
                    """)
    df_product_countries = pd.read_sql_query(query, connection, params={'countryId': country_id})
    df_product_countries['product_id'] = pd.to_numeric(df_product_countries['product_id'])

    # Read Excel file
    df = pd.read_excel(workDir + os.path.basename(f))
    df.columns = ['centrale_id', 'code_produit', 'produit_id_ancien', 'produit_id_nouveau']
    df['code_produit'] = df['code_produit'].astype(str)

    # Add central code IDs
    query_cp = text("""
                SELECT id as centrale_produit_id, centrale_id, code_produit, produit_id
                FROM centrale_produit
                WHERE country_id = :countryId""")
    df_cp = pd.read_sql_query(query_cp, connection, params={"countryId": country_id})
    df = pd.merge(df, df_cp, on=['centrale_id', 'code_produit'], how='left')

    for index, row in df.iterrows():
        cent_id = row['centrale_id']
        product_code = row['code_produit']
        old_product = row['produit_id_ancien']
        new_product = row['produit_id_nouveau']
        cent_prod_id = row['centrale_produit_id']
        if row['produit_id'] != old_product:
            df_logs = df_logs.append({'centrale_id': cent_id, 'code_produit': product_code,
                                      'produit_id_ancien': old_product, 'produit_id_nouveau': new_product,
                                      'commentaire': 'Incohérence au niveau de l\'ancien produit (' + str(row[
                                          'produit_id']) + ' en BDD)'},
                                     ignore_index=True)
        else:
            # Transferring code to new product
            upd_product = text(""" UPDATE centrale_produit SET produit_id = :prodId WHERE id = :id """)
            res_upd_prod = connection.execute(upd_product, prodId=new_product, id=cent_prod_id)
            if res_upd_prod.rowcount == 0:
                df_logs = df_logs.append({'centrale_id': cent_id, 'code_produit': product_code,
                                          'produit_id_ancien': old_product, 'produit_id_nouveau': new_product,
                                          'commentaire': 'Erreur de la mise à jour du centrale_produit'},
                                         ignore_index=True)
                continue
            # Insert country for product
            if new_product not in df_product_countries['product_id'].drop_duplicates().tolist():
                ins_prod_country = text(
                    """ INSERT INTO product_country (product_id, country_id) VALUES (:prodId, :countryId)""")
                res_ins_prod_country = connection.execute(ins_prod_country, prodId=new_product, countryId=country_id)
                df_product_countries.append({'product_id': new_product}, ignore_index=True)
                if res_ins_prod_country.rowcount == 0:
                    df_logs = df_logs.append({'centrale_id': cent_id, 'code_produit': product_code,
                                              'produit_id_ancien': old_product, 'produit_id_nouveau': new_product,
                                              'commentaire': 'Erreur de l\'ajout du pays pour le produit'},
                                             ignore_index=True)

            # For each client : update purchases
            for i in clients:
                try:
                    # Reading parameters of database
                    params_client = config(filename='config_clients.ini', section=i)

                    # Connecting to the PostgreSQL server
                    print('Connecting to the PostgreSQL database...')
                    engine_client = create_engine(URL(**params_client), echo=False)
                    connection_client = engine_client.connect()

                except (Exception, psycopg2.Error) as error_client:
                    print(error_client)
                    print(f"Error : no section for {i}")
                    continue

                try:
                    # Searching purchases for old central product
                    if i != "vetfamily":
                        upd_purchases_ref = text(""" UPDATE achats SET produit_id = :prodId WHERE id IN (SELECT distinct a.id FROM achats a WHERE a.centrale_produit_id = :centProdId)""")
                        upd_purchases = None
                    else:
                        upd_purchases_ref = text(
                            """ UPDATE ed_purchases_ref SET purr_product_id = :prodId WHERE purr_id IN (select distinct purr_id FROM ed_purchases_ref WHERE purr_central_product_id = :centProdId)""")
                        upd_purchases = text(
                            """ UPDATE ed_purchase SET purc_product_id = :prodId WHERE purc_purchase_ref_id IN (select distinct purr_id FROM ed_purchases_ref WHERE purr_central_product_id = :centProdId)""")

                    res = connection_client.execute(upd_purchases_ref, prodId=new_product, centProdId=cent_prod_id)
                    clients[i] = res.rowcount
                    if upd_purchases is not None:
                        connection_client.execute(upd_purchases, prodId=new_product, centProdId=cent_prod_id)

                except (Exception, psycopg2.Error) as error_client:
                    if connection_client:
                        print(f"Error {i} : ", error_client)

                finally:
                    # closing database connection.
                    if connection_client:
                        connection_client.close()
                        print(f"PostgreSQL connection {i} is closed")

            if country_id == constant.COUNTRY_FRANCE_ID:
                df_logs = df_logs.append(
                    {'centrale_id': cent_id, 'code_produit': product_code, 'produit_id_ancien': old_product,
                     'produit_id_nouveau': new_product, 'commentaire': 'OK', 'bourgelat': clients['bourgelat'],
                     'vetoavenir': clients['vetoavenir'], 'vetapro': clients['vetapro'],
                     'vetapharma': clients['vetapharma'], 'vetharmonie': clients['vetharmonie'],
                     'cristal': clients['cristal'], 'symbioveto': clients['symbioveto'], 'clubvet': clients['clubvet'],
                     'vetodistribution': clients['vetodistribution'], 'vetfamily': clients['vetfamily']},
                    ignore_index=True)
            else:
                df_logs = df_logs.append(
                    {'centrale_id': cent_id, 'code_produit': product_code, 'produit_id_ancien': old_product,
                     'produit_id_nouveau': new_product, 'commentaire': 'OK', 'vetfamily': clients['vetfamily']},
                    ignore_index=True)


def create_excel_file(filename, df, append):
    if append:
        wb = load_workbook(filename)
    else:
        wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=(False if append else True)):
        ws.append(r)
    wb.save(filename)


if __name__ == "__main__":
    try:
        # Reading parameters of database
        params = config(section="postgresql")

        # Connecting to the PostgreSQL server
        print('Connecting to the PostgreSQL database...')
        engine = create_engine(URL(**params), echo=False)
        connection = engine.connect()

        print(f'** Processing new products **')

        # Getting args
        args = getArguments()
        country_id = int(args.country)
        country_name = common.get_name_of_country(connection, country_id)

        now = datetime.now().strftime('%Y%m%d')

        initDir = './fichiers/doublons/' + country_name + "/"
        workDir = './encours/doublons/' + country_name + "/"
        historicDir = './historiques/doublons/' + country_name + "/" + now + "/"
        logDir = './logs/doublons/' + country_name + "/" + now + "/"

        # Create directories if not exist
        os.makedirs(workDir, exist_ok=True)
        os.makedirs(historicDir, exist_ok=True)
        os.makedirs(logDir, exist_ok=True)

        # Create empty logs dataframes
        if country_id == constant.COUNTRY_FRANCE_ID:
            df_logs = pd.DataFrame(columns=['centrale_id', 'code_produit', 'produit_id_ancien', 'produit_id_nouveau',
                                            'commentaire', 'bourgelat', 'vetoavenir', 'vetapro', 'vetapharma',
                                            'vetharmonie', 'cristal', 'symbioveto', 'clubvet', 'vetodistribution',
                                            'vetfamily'])
        else:
            df_logs = pd.DataFrame(columns=['centrale_id', 'code_produit', 'produit_id_ancien', 'produit_id_nouveau',
                                            'commentaire', 'vetfamily'])

        for f in glob.glob(r'' + initDir + F'/*.[xX][lL][sS][xX]'):
            print(f'Processing file "{os.path.basename(f)}" ...')

            # Begin transaction
            trans = connection.begin()

            # Move file to working directory
            shutil.move(f, workDir + os.path.basename(f))

            if os.stat(workDir + os.path.basename(f)).st_size > 0:
                # Process file
                process()
            else:
                print(f'File "{os.path.basename(f)}" is empty !')

            # Move source file
            shutil.move(workDir + os.path.basename(f), historicDir + os.path.basename(f))

            trans.commit()

        # Create Excel file of products errors
        if len(df_logs.index) > 0:
            os.makedirs(logDir, exist_ok=True)
            create_excel_file(logDir + "duplicate_products_logs.xlsx", df_logs.drop_duplicates(),
                              os.path.isfile(logDir + "duplicate_products_logs.xlsx"))

    except (Exception, psycopg2.Error) as error:
        if connection:
            print("Error : ", error)

    finally:
        # closing database connection.
        if connection:
            connection.close()
            print("PostgreSQL connection is closed")
